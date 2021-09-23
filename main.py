#!/usr/bin/env python

import os
import re
import sys
import logging
import argparse
import datetime
import json
import io
import zipfile

import openpyxl
import openpyxl.utils
import openpyxl.styles
import openpyxl.styles.colors
import openpyxl.writer.excel
from openpyxl.comments import Comment

import neil_tools
from neil_tools import spreadsheet_tools

import config as config_static

import arc_o365
import O365



NOW = datetime.datetime.now().astimezone()
DATESTAMP = NOW.strftime("%Y-%m-%d")
TIMESTAMP = NOW.strftime("%Y-%m-%d %H:%M:%S %Z")
FILESTAMP = NOW.strftime("%Y-%m-%d %H-%M-%S %Z")
EMAILSTAMP = NOW.strftime("%Y-%m-%d %H-%M")


def main():
    args = parse_args()
    if args.debug:
        logging.getLogger().setLevel(logging.DEBUG)
    log.debug("running...")

    config = neil_tools.init_config(config_static, ".env")

    account_avis = init_o365(config, config.TOKEN_FILENAME_AVIS)
    log.debug(f"after initializing account")

    sharepoint = account_avis.sharepoint()

    source_site = sharepoint.get_site(config.SHAREPOINT_SITE, config.SOURCE_SITE)
    source_drive = source_site.get_default_document_library(request_drive=True)

    log.debug(f"source_drive { source_drive }")

    source_folder = get_folder(source_drive, config.SOURCE_PATH)

    log.debug(f"source_folder { source_folder }")


    dest_site = sharepoint.get_site(config.SHAREPOINT_SITE, config.DEST_SITE)
    dest_drive = dest_site.get_default_document_library(request_drive=True)
    dest_folder = get_folder(dest_drive, config.DEST_PATH)

    process_source(source_folder, dest_folder)

def process_source(folder, dest_folder):

    items = folder.get_items()

    avis_regex = re.compile(r'^FY\d\d Avis Reports?')

    for item in items:
        name = item.name
        size = item.size

        is_avis_folder = (avis_regex.match(name) is not None)

        #log.debug(f"item name { name } size { size } is_avis_folder { is_avis_folder }")

        if is_avis_folder:
            process_avis_folder(item, dest_folder)

def open_child(folder, child_name):

    drive = folder.drive
    parent_path = folder.parent_path
    name = folder.name

    new_path = "/".join( [ parent_path, name, child_name ] )

    sub_folder = drive.get_item_by_path(new_path)

    return sub_folder

def process_avis_folder(source_folder, dest_parent):
    folder_name = source_folder.name

    query = dest_parent.new_query().on_attribute('name').equals(folder_name)
    
    dest_children = dest_parent.get_items(query=query)
    #log.debug(f"dest_children { dest_children }")
    dest_folder = None
    for f in dest_children:
        dest_folder = f

    if dest_folder is None:
        dest_folder = dest_parent.create_child_folder(folder_name)

    log.debug(f"dest_folder { dest_folder }")

    # cache all the file names
    dest_cache = {}
    for c in dest_folder.get_items():
        name = c.name
        dest_cache[name] = c

    items = source_folder.get_items()

    for item in items:
        name = item.name
        size = item.size

        #log.debug(f"checking item { name }")

        # only process if the name isn't already in the destination folder
        if name not in dest_cache:
            log.debug(f"copying '{ name }'")

            if name.endswith('.xlsx') and not folder_name.startswith("FY19") and not folder_name.startswith("FY20"):

                # process the spreadsheet
                bytesio = io.BytesIO()
                if not item.download(output=bytesio):
                    log.error(f"download of { name } failed")
                else:

                    try:
                        wb = openpyxl.load_workbook(bytesio)
                        bytesio.close()
                        #log.debug(f"sheetnames { wb.sheetnames }")

                        modify_avis(wb)

                        bytesio = io.BytesIO()
                        zipbuffer = zipfile.ZipFile(bytesio, mode='w')
                        writer = openpyxl.writer.excel.ExcelWriter(wb, zipbuffer)
                        writer.save()

                        # get the file size
                        buffer_data = bytesio.getvalue()
                        buffer_size = len(buffer_data)
                        #log.debug(f"buffer size is { buffer_size }")

                        # go back to the beginning
                        bytesio.seek(io.SEEK_SET, 0)

                        # upload the file into sharepoint
                        dest_folder.upload_file(None, item_name=name, stream=bytesio, stream_size=buffer_size, upload_in_chunks=False, conflict_handling="fail")
                    except:
                        # something went wrong; just copy the file
                        log.info(f"transforming file { name } failed; copying instead")
                        item.copy(target=dest_folder, name=name)
            else:
                # just copy untouched
                item.copy(target=dest_folder, name=name)

        # end of test for name being in output folder already
    # end if item loop




def modify_avis(wb):
    

    sheet_names = wb.sheetnames

    for sheet_name in sheet_names:
        modify_avis_sheet(wb[sheet_name], sheet_name)


def modify_avis_sheet(ws, sheet_name):

    table_name = sheet_name.replace(' ', '_')

    if sheet_name == 'Open RA':
        start = '$B$3'
        freeze = 'B4'
    elif sheet_name == 'Closed RA':
        start = '$A$6'
        freeze = 'A7'
    else:
        log.error(f"modify_avis_sheet: unknown sheet name '{ sheet_name }': can't process")
        return

    ws.freeze_panes = freeze

    max_rows = ws.max_row
    max_cols = ws.max_column
    last_col_letter = openpyxl.utils.get_column_letter(max_cols)
    table_ref = f"{ start }:${ last_col_letter }${ max_rows }"

    #log.debug(f"table_ref is '{ table_ref }'")
    table = openpyxl.worksheet.table.Table(displayName=table_name, ref=table_ref)
    ws.add_table(table)








def get_folder(drive, path):

    folder = drive.get_item_by_path(path)
    log.debug(f"folder { folder } is_folder { folder.is_folder }")

    return folder


def init_o365(config, token_filename=None):
    """ do initial setup to get a handle on office 365 graph api """

    if token_filename != None:
        o365 = arc_o365.arc_o365.arc_o365(config, token_filename=token_filename)
    else:
        o365 = arc_o365.arc_o365.arc_o365(config)

    account = o365.get_account()
    if account is None:
        raise Exception("could not access office 365 graph api")

    return account

    

def parse_args():
    parser = argparse.ArgumentParser(
            description="tools to support Disaster Transportation Tools reporting",
            allow_abbrev=False)
    parser.add_argument("--debug", help="turn on debugging output", action="store_true")
    parser.add_argument("-s", "--store", help="Store file on server", action="store_true")

    args = parser.parse_args()

    return args


if __name__ == "__main__":
    neil_tools.init_logging(__name__)
    log = logging.getLogger(__name__)
    main()
else:
    log = logging.getLogger(__name__)

